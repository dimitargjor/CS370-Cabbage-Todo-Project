import os, sys
from os.path import expanduser
from openpyxl import Workbook, load_workbook
from os.path import exists
from openpyxl.styles import Font
import datetime
import ColorSelection

#Returns path to ToDo List directory, and creates it if it doesn't exist
def getDirectory():    

    osType = sys.platform

    homeDir = expanduser("~")

    #If Operating system is not windows
    if (osType != 'win32'):

        #Find path to Todo-List directory
        todoPath = os.path.join(homeDir, 'Todo-List/')

        #If Todo-List Directory doesn't exist
        if os.path.isdir(todoPath) == False:

            #Make ToDo List directory
            os.mkdir(todoPath)
            ColorSelection.prRed("NOTICE: To-Do List directory created in Home directory.")

        #Return ToDo List directory
        return todoPath

    else:

        #Join home directory path to Documents
        docPath = os.path.join(homeDir, "Documents\\")

        #IF the ToDo List directory doesn't exist
        if os.path.isdir(docPath + "ToDo List\\") == False:

            #Make ToDo List directory
            os.mkdir(docPath + "ToDo List\\")
            ColorSelection.prRed("\nNOTICE: Directory Created in 'Documents'\n")

        #Make variable with ToDo List directory
        fullPath = os.path.join(docPath, "ToDo List\\")

        #Return ToDo List directory
        return fullPath

#Returns workbook for writing in excel file. Creates a new worksheet if it doesn't exist
def getWorkbook():

    todoDir = getDirectory()

    #Instance variables
    dateColumn = "A"
    prioColumn = "B"
    statColumn = "C"
    catColumn = "D"
    descColumn = "E"

    if(exists(todoDir + 'Notes.xlsx')):
        
        workbook = load_workbook(todoDir + 'Notes.xlsx')
        #worksheet =  workbook.active

    else:

        #Create a new workbook
        workbook = Workbook()
        worksheet =  workbook.active
        worksheet.title = "ToDo List"

        #Set column length
        worksheet.column_dimensions[dateColumn].width = 12
        worksheet.column_dimensions[prioColumn].width = 10
        worksheet.column_dimensions[statColumn].width = 10
        worksheet.column_dimensions[catColumn].width = 15        
        worksheet.column_dimensions[descColumn].width = 50

        #Set headings style
        worksheet[descColumn + "1"].font = Font(color = "00800080", bold = True)
        worksheet[prioColumn + "1"].font = Font(color = "00800080", bold = True) 
        worksheet[dateColumn + "1"].font = Font(color = "00800080", bold = True)
        worksheet[catColumn + "1"].font = Font(color = "00800080", bold = True) 
        worksheet[statColumn + "1"].font = Font(color = "00800080", bold = True) 

        #Set column headings
        worksheet[descColumn + "1"] = "Description"
        worksheet[prioColumn + "1"] = "Priority"
        worksheet[dateColumn + "1"] = "Due Date"
        worksheet[statColumn + "1"] = "Status"
        worksheet[catColumn + "1"] = "Category"

        workbook.save(todoDir + 'Notes.xlsx')

    return workbook

def checkExistence():

    todoDir = getDirectory()
    
    output = False

    if(exists(todoDir + 'Notes.xlsx')):
        output = True

    return output


#Function that checks if a valid date was entered
def checkDate(dateInput):
    
    output = True

    try:
        month, day, year = dateInput.split('/')
        try:
            hello = datetime.datetime(int(year), int(month), int(day))

        except ValueError:
            output = False
    except ValueError:
        output = False

    if(output == True and len(year) < 4):
        output = False

    return output

def checkCurrDate(workbook):

    from datetime import date, datetime


    todoDir = getDirectory()
    worksheet = workbook.active
    statColumn = "c"
    overdue = '\u2613'
    incomplete = u'\u25CB'

    todayDate = date.today()
    todayString = todayDate.strftime("%m/%d/%Y")
    today = datetime.strptime(todayString, "%m/%d/%Y")

    for cell in worksheet["A"]:

        currRow = cell.row

        if currRow > 1 and cell.value != 'None':

            parsedDate = datetime.strptime(cell.value, "%m/%d/%Y")

            if(worksheet[statColumn + str(currRow)] == incomplete and parsedDate < today):

                worksheet[statColumn + str(currRow)] = overdue
                workbook.save(todoDir + 'Notes.xlsx')

def getCategories():

    output = ["Analysis",
              "Design",
              "Coding",
              "Algorithms",
              "Testing" ,
              "Debugging",
              "Troubleshooting",
              "Building", 
              "Deploying",
              "Other"]

    return output

def displayCategories():

    categories = getCategories()

    ColorSelection.prGreen("Input a Category:")
    index = 1
    for category in categories:
        if(index < 10):
            ColorSelection.prPurple(" " + str(index) + ". " + category)
        else:
            ColorSelection.prPurple(str(index) + ". " + category)
        index+=1

def cls():  
    os.system('cls' if os.name == 'nt' else 'clear')
