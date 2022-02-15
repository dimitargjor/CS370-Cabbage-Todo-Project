from calendar import c
import datetime
import PrintColors
import DirControl
from TerminalControls import Clear
import openpyxl
from openpyxl import Workbook, load_workbook
from os.path import exists
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import os 

ListsDir = DirControl.FileSystem()

#Instance variables
descColumn = "A"
prioColumn = "B"
dateColumn = "C"

if(exists(ListsDir + 'Notes.xlsx')):
    workbook = load_workbook(ListsDir + 'Notes.xlsx')
    worksheet =  workbook.active

else:

    #Create a new workbook
    workbook = Workbook()
    worksheet =  workbook.active
    worksheet.title = "ToDo List"

     #Set column length
    worksheet.column_dimensions[descColumn].width = 20
    worksheet.column_dimensions[prioColumn].width = 15
    worksheet.column_dimensions[dateColumn].width = 15

    #Set headings style
    worksheet[descColumn + "1"].font = Font(color = "00800080", bold = True)
    worksheet[prioColumn + "1"].font = Font(color = "00800080", bold = True) 
    worksheet[dateColumn + "1"].font = Font(color = "00800080", bold = True) 

    #Set column headings
    worksheet[descColumn + "1"] = "Description"
    worksheet[prioColumn + "1"] = "Priority"
    worksheet[dateColumn + "1"] = "Due Date"

    workbook.save(ListsDir + 'Notes.xlsx')

def ArgNoteInput(listNotes, listPriority, listDate):

    #instance variables
    currRow = worksheet.max_row + 1

    #Assign To-Do List item description to appropriate variable
    description = listNotes

    #Assign priority to appropriate variable
    priority = listPriority

    #Check if valid priority was entered
    priority = priority.lower()
    while priority not in ("low", "medium", "high"):
        PrintColors.PrintColor.prRed("ERROR: Input Must Be High, Meduim, or Low")
        PrintColors.PrintColor.prCyan("Input the priority level (LOW/MEDIUM/HIGH):")
        priority = input()
        priority = priority.lower()

    #Assign due date to appropriate variable
    date = listDate

    #Check if entered date is valid
    if(date != None):
        while(not checkDate(date)):
            PrintColors.PrintColor.prRed("ERROR: Date Entered Is Invalid")
            PrintColors.PrintColor.prCyan("Input a due date (MM/DD/YYYY):")
            date = input()
    else:
        date = "None"

    #Write the todo list item to excel spreadsheet
    worksheet[descColumn + str(currRow)] = description
    worksheet[prioColumn + str(currRow)] = priority
    worksheet[dateColumn + str(currRow)] = date

    #Save the excel document
    workbook.save(ListsDir + 'Notes.xlsx')

def read():
    Clear.cls()
    first = True
    if (exists(ListsDir + 'Notes.xlsx')):
        spaceString = ""
        for row in worksheet.iter_rows():
            for cell in row:
                space = 25 - len(str(cell.value))
        
                for i in range(space):
                    spaceString += " "

                print (cell.value, end = spaceString)

                spaceString = ""
            print("")

            if first == True:
                print("")
                first = False

    else:
        PrintColors.PrintColor.prRed("ERROR: File Does Not Exist, You Do Not Have A To-Do List")

def readHigh():
    Clear.cls()
    if (exists(ListsDir + 'Notes.xlsx')):
        spaceString = ""

        print("Description              Priority                 Due Date" + "\n")

        for row in worksheet.iter_rows():

            if row[1].value == 'High':
                for cell in row:
                    space = 25 - len(str(cell.value))
                    for i in range(space):
                        spaceString += " "
                    print (cell.value, end = spaceString)
                    
                    spaceString = ""
                print("")
    else:
        PrintColors.PrintColor.prRed("ERROR: File Does Not Exist, You Do Not Have A To-Do List")

def clear():
    Clear.cls()
    PrintColors.PrintColor.prRed("***WARNING***")
    PrintColors.PrintColor.prGreen("Are you sure you want to clear your To-Do List?")
    PrintColors.PrintColor.prGreen("To clear list, type 'CONFIRM'. Type any other key to cancel:")
    choice = input(" Your Input: ")
    choice = choice.lower()
    if (choice == "confirm"):
        Clear.cls()
        worksheet.delete_rows(2, worksheet.max_row+1)

        workbook.save(ListsDir + 'Notes.xlsx')
        Clear.cls()
        PrintColors.PrintColor.prPurple("You have cleared your file.")

    else:
        Clear.cls()
        PrintColors.PrintColor.prPurple("File clear has been cancelled.")

def addNotes():

    #instance variables
    currRow = worksheet.max_row + 1

    Clear.cls()
    
    #Input todo list item description
    PrintColors.PrintColor.prCyan("Input the description of the To-Do List item:")
    description = input(" Your Input: ")

    Clear.cls()

    #Input priority level
    PrintColors.PrintColor.prCyan("Input priority level:")
    PrintColors.PrintColor.prYellow("1. Low  \n " +
        "2. Medium  \n " +
        "3. High")
    priority = input(" Your Input: ")

    #Check if valid priority was entered
    while priority not in ("1", "2", "3"):
        PrintColors.PrintColor.prRed("ERROR: Input Must Be High, Meduim, or Low")
        PrintColors.PrintColor.prCyan("Input priority level:")
        PrintColors.PrintColor.prYellow("1. Low  \n " +
            "2. Medium  \n " +
            "3. High")
        priority = input(" Your Input: ")

    if(priority == '1'):
        priority = "Low"
    elif(priority == '2'):
        priority = "Medium"
    elif(priority == '3'):
        priority = "High"

    Clear.cls()
    
    #Input due date
    PrintColors.PrintColor.prCyan("Input a due date (MM/DD/YYYY) or '-1' to skip:")
    date = input(" Your Input: ")

    #Check if entered date is valid
    if(date != '-1'):
        while(not checkDate(date)):
            PrintColors.PrintColor.prRed("ERROR: Date Entered Is Invalid")
            PrintColors.PrintColor.prCyan("Input a due date (MM/DD/YYYY):")
            date = input(" Your Input: ")
    else:
        date = "None"
    #Write the todo list item to excel spreadsheet
    worksheet[descColumn + str(currRow)] = description
    worksheet[prioColumn + str(currRow)] = priority
    worksheet[dateColumn + str(currRow)] = date

    #Save the excel document
    workbook.save(ListsDir + 'Notes.xlsx')

#Function that checks if a valid date was entered
def checkDate(dateInput):

    output = True

    try:
        month, day, year = dateInput.split('/')
        try:
            datetime.datetime(int(year), int(month), int(day))
        except ValueError:
            output = False
    except ValueError:
        output = False

    return output
